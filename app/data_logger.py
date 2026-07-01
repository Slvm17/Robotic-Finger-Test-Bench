"""
Data logger for LoadCellTester.
Accumulates completed test records and exports them to CSV or Excel.
Only peak force is ever stored — never instantaneous readings.
"""
import csv
import os
from datetime import datetime
from typing import Any, Dict, List, Optional


try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


FIELDS = ["test_id", "operator", "date", "time", "sample_name", "peak_force_N"]


class DataLogger:
    """Stores test results in memory and writes them to CSV / Excel on demand."""

    def __init__(self, export_folder: str = "./exports") -> None:
        self._export_folder = export_folder
        self._records: List[Dict[str, Any]] = []

    # ── Configuration ─────────────────────────────────────────────────────────

    def set_export_folder(self, folder: str) -> None:
        self._export_folder = folder

    # ── Record management ─────────────────────────────────────────────────────

    def add_record(
        self,
        test_id: str,
        operator: str,
        sample_name: str,
        peak_force: float,
        timestamp: Optional[datetime] = None,
    ) -> None:
        """Append one completed test record.  Only peak_force is stored."""
        ts = timestamp or datetime.now()
        self._records.append(
            {
                "test_id": test_id,
                "operator": operator,
                "date": ts.strftime("%Y-%m-%d"),
                "time": ts.strftime("%H:%M:%S"),
                "sample_name": sample_name,
                "peak_force_N": round(peak_force, 4),
            }
        )

    def get_records(self) -> List[Dict[str, Any]]:
        return list(self._records)

    def clear(self) -> None:
        self._records = []

    # ── Export ────────────────────────────────────────────────────────────────

    def _ensure_folder(self) -> str:
        os.makedirs(self._export_folder, exist_ok=True)
        return self._export_folder

    def _default_path(self, ext: str) -> str:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return os.path.join(self._ensure_folder(), f"results_{ts}.{ext}")

    def export_csv(self, path: Optional[str] = None) -> str:
        """Write all records to a CSV file.  Returns the file path."""
        path = path or self._default_path("csv")
        with open(path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=FIELDS)
            writer.writeheader()
            writer.writerows(self._records)
        return path

    def export_excel(self, path: Optional[str] = None) -> str:
        """Write all records to an Excel file.  Returns the file path."""
        if not XLSX_AVAILABLE:
            raise ImportError(
                "openpyxl is not installed.\nRun:  pip install openpyxl"
            )
        path = path or self._default_path("xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Results"

        # Header row
        ws.append(FIELDS)
        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for rec in self._records:
            ws.append([rec[f] for f in FIELDS])

        # Auto column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        wb.save(path)
        return path
